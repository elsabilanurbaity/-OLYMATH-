import streamlit as st
import pandas as pd
import numpy as np
import pickle
import io
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="OLYMATH – Seleksi Olimpiade Matematika",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif; }

/* Background */
.stApp { background: linear-gradient(135deg, #0f0c29 0%, #1a1a4e 50%, #24243e 100%); }

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a1a4e 0%, #0f0c29 100%);
    border-right: 1px solid rgba(255,255,255,0.1);
}
[data-testid="stSidebar"] * { color: #e0e0ff !important; }

/* Header brand */
.brand-header {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    border-radius: 20px;
    padding: 28px 24px;
    text-align: center;
    margin-bottom: 24px;
    box-shadow: 0 8px 32px rgba(102,126,234,0.4);
}
.brand-title {
    font-size: 3rem;
    font-weight: 800;
    color: white;
    letter-spacing: 4px;
    text-shadow: 0 2px 12px rgba(0,0,0,0.3);
    margin: 0;
}
.brand-sub {
    color: rgba(255,255,255,0.85);
    font-size: 0.9rem;
    margin-top: 6px;
    letter-spacing: 1px;
}

/* Cards */
.card {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 16px;
    padding: 24px;
    margin-bottom: 20px;
    backdrop-filter: blur(10px);
}
.card-title {
    font-size: 1.1rem;
    font-weight: 700;
    color: #a78bfa;
    margin-bottom: 16px;
    display: flex;
    align-items: center;
    gap: 8px;
}

/* Result cards */
.result-siap {
    background: linear-gradient(135deg, rgba(16,185,129,0.15), rgba(5,150,105,0.1));
    border: 2px solid rgba(16,185,129,0.5);
    border-radius: 20px; padding: 28px; text-align: center;
}
.result-potensial {
    background: linear-gradient(135deg, rgba(245,158,11,0.15), rgba(217,119,6,0.1));
    border: 2px solid rgba(245,158,11,0.5);
    border-radius: 20px; padding: 28px; text-align: center;
}
.result-tidak {
    background: linear-gradient(135deg, rgba(239,68,68,0.15), rgba(185,28,28,0.1));
    border: 2px solid rgba(239,68,68,0.5);
    border-radius: 20px; padding: 28px; text-align: center;
}
.result-label {
    font-size: 2rem; font-weight: 800; margin: 8px 0;
}
.result-emoji { font-size: 3rem; }

/* Score bar */
.score-bar-bg {
    background: rgba(255,255,255,0.1);
    border-radius: 999px; height: 10px; margin: 6px 0;
}
.score-bar-fill {
    height: 10px; border-radius: 999px;
    background: linear-gradient(90deg, #667eea, #764ba2);
}

/* Metric box */
.metric-box {
    background: rgba(102,126,234,0.1);
    border: 1px solid rgba(102,126,234,0.3);
    border-radius: 12px; padding: 16px; text-align: center;
}
.metric-val { font-size: 1.8rem; font-weight: 800; color: #a78bfa; }
.metric-lbl { font-size: 0.75rem; color: rgba(255,255,255,0.6); margin-top: 4px; }

/* Info box */
.info-box {
    background: rgba(102,126,234,0.08);
    border-left: 4px solid #667eea;
    border-radius: 0 12px 12px 0;
    padding: 14px 18px;
    margin: 12px 0;
    color: rgba(255,255,255,0.85);
    font-size: 0.9rem;
}

/* Template download button area */
.template-box {
    background: linear-gradient(135deg, rgba(102,126,234,0.12), rgba(118,75,162,0.12));
    border: 1px dashed rgba(102,126,234,0.5);
    border-radius: 16px;
    padding: 20px 24px;
    margin-bottom: 20px;
    text-align: center;
}

/* All text white on dark bg */
p, li, span, label, .stMarkdown { color: rgba(255,255,255,0.85) !important; }
h1,h2,h3,h4 { color: white !important; }
.stTextInput label, .stNumberInput label, .stSelectbox label { color: rgba(255,255,255,0.8) !important; }

/* Input fields */
.stNumberInput input, .stTextInput input {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    border-radius: 10px !important; color: white !important;
}
.stNumberInput input:focus, .stTextInput input:focus {
    border-color: #667eea !important;
    box-shadow: 0 0 0 2px rgba(102,126,234,0.3) !important;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: rgba(255,255,255,0.05);
    border-radius: 12px; padding: 4px; gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px; color: rgba(255,255,255,0.6) !important;
    font-weight: 500;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #667eea, #764ba2) !important;
    color: white !important;
}

/* Buttons */
.stButton button {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
    color: white !important; border: none !important;
    border-radius: 12px !important; font-weight: 600 !important;
    padding: 10px 24px !important;
    box-shadow: 0 4px 15px rgba(102,126,234,0.4) !important;
    transition: all 0.2s !important;
}
.stButton button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(102,126,234,0.5) !important;
}

/* Download button */
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #10b981, #059669) !important;
    box-shadow: 0 4px 15px rgba(16,185,129,0.4) !important;
}

/* DataFrame */
.stDataFrame { border-radius: 12px; overflow: hidden; }
[data-testid="stDataFrameResizable"] { border-radius: 12px; }

/* Divider */
hr { border-color: rgba(255,255,255,0.1) !important; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.05) !important;
    border: 2px dashed rgba(102,126,234,0.4) !important;
    border-radius: 16px !important;
}

/* Success/warning/error */
.stSuccess { background: rgba(16,185,129,0.1) !important; border-radius: 12px !important; }
.stWarning { background: rgba(245,158,11,0.1) !important; border-radius: 12px !important; }
.stError { background: rgba(239,68,68,0.1) !important; border-radius: 12px !important; }

/* Expander */
.streamlit-expanderHeader {
    background: rgba(255,255,255,0.05) !important;
    border-radius: 12px !important; color: white !important;
}
</style>
""", unsafe_allow_html=True)

# ── Load model ────────────────────────────────────────────────────────────────
@st.cache_resource
def load_model():
    model_path = os.path.join(os.path.dirname(__file__), "model.pkl")
    with open(model_path, "rb") as f:
        data = pickle.load(f)
    return data["model"], data["scaler"], data["features"]

try:
    model, scaler, features = load_model()
except Exception as e:
    st.error(f"❌ Gagal memuat model: {e}")
    st.stop()

# ── Constants ─────────────────────────────────────────────────────────────────
FEATURE_LABELS = {
    "NUM_ALJ": "Numerasi Aljabar",
    "NUM_GEO": "Numerasi Geometri",
    "NUM_BIL": "Numerasi Bilangan",
    "NUM_DAT": "Data & Ketidakpastian",
    "NUM_L3": "Kemampuan Menalar",
    "LIT":     "Literasi Matematika",
}
FEATURE_ICONS = {
    "NUM_ALJ": "𝑥²",
    "NUM_GEO": "📐",
    "NUM_BIL": "🔢",
    "NUM_DAT": "📊",
    "NUM_L3":  "🧠",
    "LIT":     "📚",
}

# ── Helper functions ──────────────────────────────────────────────────────────
def hitung_skor_kesiapan(values):
    """Skor kesiapan komposit 0-100 yang selalu monoton naik seiring nilai naik."""
    arr = np.array(values, dtype=float)
    rata = np.mean(arr)
    min_val = np.min(arr)
    # 70% rata-rata + 30% nilai minimum (untuk mempertimbangkan konsistensi)
    return round(0.7 * rata + 0.3 * min_val, 1)

def predict_single(nilai_list):
    arr = np.array([nilai_list])
    arr_scaled = scaler.transform(pd.DataFrame(arr, columns=features))
    status = model.predict(arr_scaled)[0]
    proba = model.predict_proba(arr_scaled)[0]
    classes = model.classes_
    proba_dict = dict(zip(classes, proba))
    # Confidence = probabilitas kelas yang diprediksi
    confidence = proba_dict[status] * 100
    skor_kesiapan = hitung_skor_kesiapan(nilai_list)
    return status, confidence, proba_dict, skor_kesiapan

def status_emoji(status):
    return {"Siap Olimpiade": "🏆", "Potensial": "⭐", "Tidak Siap": "📚"}.get(status, "❓")

def status_color(status):
    return {"Siap Olimpiade": "#10b981", "Potensial": "#f59e0b", "Tidak Siap": "#ef4444"}.get(status, "#888")

def status_css_class(status):
    return {"Siap Olimpiade": "result-siap", "Potensial": "result-potensial", "Tidak Siap": "result-tidak"}.get(status, "card")

def generate_review_rekomendasi(nama, nilai_dict, status, skor_kesiapan):
    rata = np.mean(list(nilai_dict.values()))
    sorted_feats = sorted(nilai_dict.items(), key=lambda x: x[1])
    lemah = sorted_feats[:2]
    kuat = sorted_feats[-2:]

    kuat_str = " dan ".join([f"{FEATURE_LABELS[k]} ({v:.1f})" for k, v in kuat])
    lemah_str = " dan ".join([f"{FEATURE_LABELS[k]} ({v:.1f})" for k, v in lemah])

    if status == "Siap Olimpiade":
        review = (
            f"{nama} menunjukkan performa luar biasa dengan rata-rata {rata:.1f} dan "
            f"skor kesiapan {skor_kesiapan}. Siswa unggul terutama di {kuat_str}. "
            f"Dengan konsistensi nilai yang tinggi, siswa ini sangat layak mengikuti seleksi olimpiade matematika."
        )
        rekomendasi = [
            "✅ Daftarkan ke seleksi olimpiade matematika tingkat berikutnya",
            "📘 Perdalam soal-soal olimpiade level nasional/internasional (OSN, IMO)",
            f"🎯 Pertahankan keunggulan di {FEATURE_LABELS[kuat[-1][0]]} dan perkuat area {FEATURE_LABELS[lemah[0][0]]}",
            "🤝 Ikutsertakan dalam kelompok belajar kompetitif / coaching olimpiade",
            "📈 Latihan soal Higher Order Thinking Skills (HOTS) secara rutin",
        ]
        saran = (
            f"Saran utama: Fokus pada latihan soal tipe olimpiade yang membutuhkan "
            f"penalaran mendalam, terutama pada topik {FEATURE_LABELS[lemah[0][0]]} yang masih bisa ditingkatkan."
        )
    elif status == "Potensial":
        review = (
            f"{nama} memiliki potensi yang baik dengan rata-rata {rata:.1f} dan skor kesiapan {skor_kesiapan}. "
            f"Kekuatan terlihat pada {kuat_str}. Namun, perlu penguatan pada {lemah_str} "
            f"agar dapat bersaing di olimpiade matematika."
        )
        rekomendasi = [
            f"📚 Intensifkan latihan pada {FEATURE_LABELS[lemah[0][0]]} dan {FEATURE_LABELS[lemah[1][0]]}",
            "🧩 Ikuti bimbingan belajar atau ekstrakurikuler matematika intensif",
            "📝 Kerjakan soal olimpiade tingkat kabupaten/kota secara rutin",
            "⏱️ Evaluasi ulang dalam 1-2 bulan setelah intervensi pembelajaran",
            "💡 Tingkatkan kemampuan problem-solving lintas topik matematika",
        ]
        saran = (
            f"Saran utama: Dengan perbaikan pada {FEATURE_LABELS[lemah[0][0]]} (skor {lemah[0][1]:.1f}), "
            f"siswa ini berpotensi masuk kategori Siap Olimpiade. Berikan latihan terfokus dan motivasi."
        )
    else:
        review = (
            f"{nama} saat ini belum siap untuk olimpiade matematika, dengan rata-rata {rata:.1f} dan "
            f"skor kesiapan {skor_kesiapan}. Nilai di {lemah_str} memerlukan perhatian khusus. "
            f"Diperlukan program penguatan dasar yang terstruktur sebelum mengikuti seleksi."
        )
        rekomendasi = [
            f"🔧 Perkuat pemahaman dasar pada {FEATURE_LABELS[lemah[0][0]]} dan {FEATURE_LABELS[lemah[1][0]]}",
            "📖 Ikuti program remedial / penguatan konsep matematika dasar",
            "👩‍🏫 Dapatkan bimbingan intensif dari guru matematika",
            "🗓️ Tetapkan target belajar harian dengan soal-soal bertahap",
            "🔄 Ikuti tes ulang setelah 2-3 bulan program penguatan",
        ]
        saran = (
            f"Saran utama: Prioritaskan penguatan di {FEATURE_LABELS[lemah[0][0]]} ({lemah[0][1]:.1f}) dan "
            f"{FEATURE_LABELS[lemah[1][0]]} ({lemah[1][1]:.1f}). Jangan putus semangat – dengan kerja keras "
            f"siswa ini masih bisa berkembang pesat!"
        )

    return review, rekomendasi, saran

# ── Excel template generator ──────────────────────────────────────────────────
def buat_template_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Input"

    # Header info
    ws.merge_cells("A1:G1")
    ws["A1"] = "🏆 OLYMATH – Template Input Data Siswa"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="3730A3")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = "Isi data siswa mulai baris ke-4. Semua nilai dalam rentang 0–100."
    ws["A2"].font = Font(italic=True, size=10, color="4B5563")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    headers = ["Nama Siswa", "Numerasi Aljabar", "Numerasi Geometri",
               "Numerasi Bilangan", "Data & Ketidakpastian", "Kemampuan Menalar", "Literasi Matematika"]
    col_keys = ["A", "B", "C", "D", "E", "F", "G"]
    header_colors = ["1E40AF", "7C3AED", "2563EB", "0891B2", "059669", "D97706", "DC2626"]

    for i, (h, col, color) in enumerate(zip(headers, col_keys, header_colors), 1):
        cell = ws[f"{col}3"]
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 40

    # Sample data
    samples = [
        ["Budi Santoso", 72.5, 68.3, 75.0, 70.2, 73.5, 69.8],
        ["Siti Rahayu", 85.0, 82.1, 88.5, 79.6, 84.3, 86.2],
        ["Ahmad Fauzi", 55.2, 48.7, 62.1, 53.4, 58.0, 51.3],
    ]
    light_fill = PatternFill("solid", fgColor="EEF2FF")
    border = Border(
        left=Side(style="thin", color="C7D2FE"),
        right=Side(style="thin", color="C7D2FE"),
        top=Side(style="thin", color="C7D2FE"),
        bottom=Side(style="thin", color="C7D2FE"),
    )
    for r, row in enumerate(samples, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if r % 2 == 0:
                cell.fill = light_fill
            if c > 1:
                cell.alignment = Alignment(horizontal="center")

    # Empty rows for input
    for r in range(7, 57):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = border
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Excel result exporter ─────────────────────────────────────────────────────
def export_hasil_excel(results_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Prediksi"

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "🏆 OLYMATH – Hasil Seleksi Tahap Awal Olimpiade Matematika"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="3730A3")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Digenerate oleh OLYMATH | Total Siswa: {len(results_df)}"
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "No", "Nama Siswa", "Num. Aljabar", "Num. Geometri",
        "Num. Bilangan", "Data & Ketidakpastian", "Kemampuan Menalar",
        "Literasi", "Skor Kesiapan", "Status Kesiapan",
    ]
    header_colors = ["1E3A5F","1E3A5F","5B21B6","1D4ED8","0E7490","065F46","92400E","991B1B","374151","374151"]
    for i, (h, color) in enumerate(zip(headers, header_colors), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 36

    status_colors = {
        "Siap Olimpiade": ("D1FAE5", "065F46"),
        "Potensial":       ("FEF3C7", "92400E"),
        "Tidak Siap":      ("FEE2E2", "991B1B"),
    }
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for idx, row in results_df.iterrows():
        r = idx + 4
        status = row.get("Status Kesiapan", "")
        bg, fg = status_colors.get(status, ("F9FAFB", "111827"))

        vals = [
            idx + 1,
            row.get("Nama Siswa", f"Siswa {idx+1}"),
            row.get("NUM_ALJ", ""),
            row.get("NUM_GEO", ""),
            row.get("NUM_BIL", ""),
            row.get("NUM_DAT", ""),
            row.get("NUM_L3", ""),
            row.get("LIT", ""),
            row.get("Skor Kesiapan", ""),
            row.get("Status Kesiapan", ""),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")
            if c == 10:  # Status column
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(bold=True, color=fg, size=9)

    # Column widths
    col_widths = [5, 22, 14, 14, 14, 18, 16, 12, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Ringkasan")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Ringkasan Hasil Seleksi"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="3730A3")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    counts = results_df["Status Kesiapan"].value_counts()
    total = len(results_df)
    summary_data = [
        ("Status", "Jumlah", "Persentase"),
        ("Siap Olimpiade", counts.get("Siap Olimpiade", 0), f'=B3/{total}*100'),
        ("Potensial",       counts.get("Potensial", 0),       f'=B4/{total}*100'),
        ("Tidak Siap",      counts.get("Tidak Siap", 0),      f'=B5/{total}*100'),
        ("Total",           total,                             "100%"),
    ]
    hdr_fills = [None, "D1FAE5", "FEF3C7", "FEE2E2", "E5E7EB"]
    for r, (row_data, fill) in enumerate(zip(summary_data, hdr_fills), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if r == 2:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if c == 3 and r > 2 and r < 6:
                cell.number_format = "0.00"

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div class="brand-header">
        <div class="brand-title">OLYMATH</div>
        <div class="brand-sub">Sistem Seleksi Olimpiade Matematika</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 📌 Navigasi")
    page = st.radio(
        "", ["🏠 Beranda", "👤 Input Satu Siswa", "📂 Upload File Excel", "📖 Tentang Sistem"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown("""
    <div class="info-box">
    <b>🎯 Kriteria Status:</b><br><br>
    🏆 <b>Siap Olimpiade</b><br>
    Rata-rata ≥ 80 & semua nilai ≥ 40<br><br>
    ⭐ <b>Potensial</b><br>
    Rata-rata ≥ 65 & nilai min ≥ 35<br><br>
    📚 <b>Tidak Siap</b><br>
    Di bawah kriteria potensial
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box" style="margin-top:12px">
    <b>🤖 Model:</b> Random Forest<br>
    <b>📊 Akurasi:</b> 95.4%<br>
    <b>📁 Data:</b> 29,000+ siswa
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HALAMAN BERANDA
# ══════════════════════════════════════════════════════════════════════════════
if page == "🏠 Beranda":
    st.markdown("""
    <div style="text-align:center; padding: 40px 20px 20px;">
        <div style="font-size:5rem;">🏆</div>
        <h1 style="font-size:3.5rem; font-weight:900; background: linear-gradient(135deg,#667eea,#a78bfa,#ec4899);
            -webkit-background-clip:text; -webkit-text-fill-color:transparent; margin:10px 0;">
            OLYMATH
        </h1>
        <p style="font-size:1.2rem; color:rgba(255,255,255,0.7); max-width:600px; margin:0 auto;">
            Sistem Cerdas Seleksi Tahap Awal Calon Peserta Olimpiade Matematika<br>
            berbasis Machine Learning
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    for col, emoji, title, desc in [
        (c1, "🤖", "AI-Powered", "Menggunakan Random Forest dengan akurasi 95.4% terlatih pada 29.000+ data siswa"),
        (c2, "⚡", "Cepat & Akurat", "Prediksi instan untuk satu siswa atau ratusan siswa sekaligus via upload Excel"),
        (c3, "📋", "Laporan Lengkap", "Output berupa status, skor kesiapan, review, rekomendasi, dan saran spesifik"),
    ]:
        col.markdown(f"""
        <div class="card" style="text-align:center; height:160px;">
            <div style="font-size:2.5rem;">{emoji}</div>
            <div style="font-weight:700; color:#a78bfa; margin:8px 0;">{title}</div>
            <div style="font-size:0.85rem; color:rgba(255,255,255,0.65);">{desc}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### 📊 Komponen Penilaian")
    cols = st.columns(3)
    for i, (key, label) in enumerate(FEATURE_LABELS.items()):
        with cols[i % 3]:
            st.markdown(f"""
            <div class="metric-box" style="margin-bottom:12px;">
                <div style="font-size:1.8rem;">{FEATURE_ICONS[key]}</div>
                <div style="font-size:0.85rem; font-weight:600; color:#c4b5fd; margin-top:4px;">{label}</div>
                <div style="font-size:0.75rem; color:rgba(255,255,255,0.5);">Skala 0 – 100</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
    <div class="card">
        <div class="card-title">💡 Catatan Penting tentang Tingkat Keyakinan Model</div>
        <p style="color:rgba(255,255,255,0.8); font-size:0.9rem;">
        Sistem ini menampilkan <b>Skor Kesiapan</b> (0–100) yang selalu naik seiring meningkatnya nilai siswa — 
        skor ini merupakan indikator utama kemajuan siswa. <br><br>
        <b>Tingkat Keyakinan Model</b> adalah kepercayaan AI dalam menentukan <i>kategori</i>, bukan ukuran 
        kemampuan siswa. Siswa dengan skor mendekati batas antar-kategori (misalnya 78–82) akan memiliki 
        keyakinan lebih rendah karena berada di "zona abu-abu" — ini adalah perilaku normal algoritma klasifikasi. 
        Selalu gunakan <b>Skor Kesiapan</b> sebagai patokan utama perbandingan antar siswa.
        </p>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HALAMAN INPUT SATU SISWA
# ══════════════════════════════════════════════════════════════════════════════
elif page == "👤 Input Satu Siswa":
    st.markdown("## 👤 Prediksi Kesiapan Olimpiade – Satu Siswa")
    st.markdown("<div class='info-box'>Masukkan data nilai siswa pada form di bawah, kemudian klik <b>Analisis Sekarang</b>.</div>", unsafe_allow_html=True)

    with st.form("form_siswa"):
        nama = st.text_input("Nama Siswa", placeholder="contoh: Budi Santoso")
        st.markdown("**📝 Nilai Kompetensi (skala 0–100)**")
        c1, c2, c3 = st.columns(3)
        with c1:
            alj = st.number_input("𝑥² Numerasi Aljabar", 0.0, 100.0, 70.0, step=0.5)
            dat = st.number_input("📊 Data & Ketidakpastian", 0.0, 100.0, 70.0, step=0.5)
        with c2:
            geo = st.number_input("📐 Numerasi Geometri", 0.0, 100.0, 70.0, step=0.5)
            l3  = st.number_input("🧠 Kemampuan Menalar", 0.0, 100.0, 70.0, step=0.5)
        with c3:
            bil = st.number_input("🔢 Numerasi Bilangan", 0.0, 100.0, 70.0, step=0.5)
            lit = st.number_input("📚 Literasi Matematika", 0.0, 100.0, 70.0, step=0.5)
        submitted = st.form_submit_button("🔍 Analisis Sekarang", use_container_width=True)

    if submitted:
        if not nama.strip():
            st.warning("⚠️ Mohon isi nama siswa terlebih dahulu.")
        else:
            nilai_list = [alj, geo, bil, dat, l3, lit]
            nilai_dict = dict(zip(features, nilai_list))
            status, confidence, proba_dict, skor_kesiapan = predict_single(nilai_list)
            review, rekomendasi, saran = generate_review_rekomendasi(nama, nilai_dict, status, skor_kesiapan)

            st.markdown("---")
            st.markdown("### 🎯 Hasil Analisis")

            # Result card
            css_cls = status_css_class(status)
            color = status_color(status)
            emoji = status_emoji(status)
            st.markdown(f"""
            <div class="{css_cls}">
                <div class="result-emoji">{emoji}</div>
                <div style="color:{color}; font-size:0.9rem; font-weight:600; letter-spacing:2px; text-transform:uppercase;">Status Kesiapan</div>
                <div class="result-label" style="color:{color};">{status}</div>
                <div style="color:rgba(255,255,255,0.7); font-size:0.9rem;">{nama}</div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            m1, m2, m3 = st.columns(3)
            with m1:
                rata = np.mean(nilai_list)
                st.markdown(f"""<div class="metric-box">
                    <div class="metric-val">{rata:.1f}</div>
                    <div class="metric-lbl">Rata-rata Nilai</div>
                </div>""", unsafe_allow_html=True)
            with m2:
                st.markdown(f"""<div class="metric-box">
                    <div class="metric-val" style="color:#10b981;">{skor_kesiapan}</div>
                    <div class="metric-lbl">Skor Kesiapan ✦</div>
                </div>""", unsafe_allow_html=True)
            with m3:
                st.markdown(f"""<div class="metric-box">
                    <div class="metric-val" style="color:{color};">{confidence:.1f}%</div>
                    <div class="metric-lbl">Keyakinan Model</div>
                </div>""", unsafe_allow_html=True)

            st.caption("✦ Skor Kesiapan selalu naik seiring meningkatnya nilai — gunakan ini untuk membandingkan antar siswa dalam kategori yang sama.")

            # Profil nilai
            st.markdown("#### 📊 Profil Nilai Kompetensi")
            for key, val in nilai_dict.items():
                lbl = FEATURE_LABELS[key]
                icon = FEATURE_ICONS[key]
                bar_pct = int(val)
                color_bar = "#10b981" if val >= 80 else "#f59e0b" if val >= 65 else "#ef4444"
                st.markdown(f"""
                <div style="margin-bottom:8px;">
                    <div style="display:flex; justify-content:space-between; margin-bottom:3px;">
                        <span style="font-size:0.85rem; color:rgba(255,255,255,0.8);">{icon} {lbl}</span>
                        <span style="font-size:0.85rem; font-weight:700; color:{color_bar};">{val:.1f}</span>
                    </div>
                    <div class="score-bar-bg">
                        <div class="score-bar-fill" style="width:{bar_pct}%; background:{color_bar};"></div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            # Review & Rekomendasi
            st.markdown("#### 📝 Review")
            st.markdown(f"""<div class="card"><p>{review}</p></div>""", unsafe_allow_html=True)

            st.markdown("#### 🎯 Rekomendasi")
            rek_html = "".join([f"<li style='margin-bottom:8px; color:rgba(255,255,255,0.85);'>{r}</li>" for r in rekomendasi])
            st.markdown(f"""<div class="card"><ul style="margin:0; padding-left:20px;">{rek_html}</ul></div>""", unsafe_allow_html=True)

            st.markdown("#### 💬 Saran")
            st.markdown(f"""<div class="card"><p>{saran}</p></div>""", unsafe_allow_html=True)

            # Download
            st.markdown("#### 📥 Unduh Hasil")
            row_data = {
                "Nama Siswa": nama,
                "NUM_ALJ": alj, "NUM_GEO": geo, "NUM_BIL": bil,
                "NUM_DAT": dat, "NUM_L3": l3, "LIT": lit,
                "Skor Kesiapan": skor_kesiapan,
                "Status Kesiapan": status,
            }
            result_df = pd.DataFrame([row_data])
            excel_buf = export_hasil_excel(result_df)
            safe_name = nama.replace(" ", "_").replace("/", "-")
            st.download_button(
                "📥 Download Hasil Excel",
                data=excel_buf,
                file_name=f"OLYMATH_{safe_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ══════════════════════════════════════════════════════════════════════════════
# HALAMAN UPLOAD FILE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📂 Upload File Excel":
    st.markdown("## 📂 Upload File Excel – Banyak Siswa")

    # ── Template download ──────────────────────────────────────────────────
    st.markdown("""
    <div class="template-box">
        <div style="font-size:2rem;">📋</div>
        <div style="font-weight:700; color:#a78bfa; font-size:1.1rem; margin:8px 0;">Unduh Template Excel</div>
        <div style="color:rgba(255,255,255,0.65); font-size:0.85rem; margin-bottom:16px;">
            Belum punya file? Download template resmi OLYMATH di bawah ini.<br>
            Isi data siswa sesuai format, lalu upload kembali.
        </div>
    </div>
    """, unsafe_allow_html=True)

    template_buf = buat_template_excel()
    st.download_button(
        "📥 Download Template Excel",
        data=template_buf,
        file_name="OLYMATH_Template_Input.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )

    st.markdown("---")

    # ── File uploader ──────────────────────────────────────────────────────
    st.markdown("### 📤 Upload File Excel Siswa")
    st.markdown("""
    <div class="info-box">
    <b>Format kolom yang dibutuhkan:</b> Nama Siswa, Numerasi Aljabar, Numerasi Geometri, 
    Numerasi Bilangan, Data & Ketidakpastian, Kemampuan Menalar, Literasi Matematika<br>
    <b>Atau gunakan kode kolom:</b> NUM_ALJ, NUM_GEO, NUM_BIL, NUM_DAT, NUM_L3, LIT
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx", "xls"])

    if uploaded:
        try:
            df_input = pd.read_excel(uploaded)
            st.success(f"✅ File berhasil dibaca: **{len(df_input)} baris** terdeteksi.")

            # Normalisasi nama kolom
            rename_map = {
                "nama siswa": "Nama Siswa", "nama": "Nama Siswa",
                "numerasi aljabar": "NUM_ALJ",
                "numerasi geometri": "NUM_GEO",
                "numerasi bilangan": "NUM_BIL",
                "data & ketidakpastian": "NUM_DAT", "data dan ketidakpastian": "NUM_DAT",
                "kemampuan menalar": "NUM_L3", "menalar": "NUM_L3",
                "literasi matematika": "LIT", "literasi": "LIT",
            }
            df_input.columns = [c.strip() for c in df_input.columns]
            df_input = df_input.rename(columns={c: rename_map.get(c.lower(), c) for c in df_input.columns})

            # Cek kolom
            missing = [f for f in features if f not in df_input.columns]
            if missing:
                st.error(f"❌ Kolom berikut tidak ditemukan: {', '.join(missing)}\n\nPastikan nama kolom sesuai template.")
                st.stop()

            # Preview
            with st.expander("👁️ Preview data (5 baris pertama)"):
                st.dataframe(df_input.head(), use_container_width=True)

            if st.button("🚀 Mulai Prediksi Semua Siswa", use_container_width=True):
                progress = st.progress(0, text="Memproses data...")
                results = []
                for i, row in df_input.iterrows():
                    try:
                        vals = [float(row[f]) for f in features]
                        nama_siswa = str(row.get("Nama Siswa", f"Siswa {i+1}"))
                        status, confidence, proba_dict, skor_kesiapan = predict_single(vals)
                        review, rek, saran = generate_review_rekomendasi(nama_siswa, dict(zip(features, vals)), status, skor_kesiapan)
                        results.append({
                            "Nama Siswa": nama_siswa,
                            "NUM_ALJ": vals[0], "NUM_GEO": vals[1], "NUM_BIL": vals[2],
                            "NUM_DAT": vals[3], "NUM_L3": vals[4], "LIT": vals[5],
                            "Skor Kesiapan": skor_kesiapan,
                            "Status Kesiapan": status,
                            "Keyakinan Model (%)": round(confidence, 1),
                            "Review": review,
                            "Saran": saran,
                        })
                    except Exception:
                        pass
                    progress.progress((i + 1) / len(df_input), text=f"Memproses {i+1}/{len(df_input)} siswa...")

                progress.empty()
                results_df = pd.DataFrame(results)
                st.session_state["results_df"] = results_df
                st.success(f"✅ Selesai! {len(results_df)} siswa berhasil dianalisis.")

        except Exception as e:
            st.error(f"❌ Gagal membaca file: {e}")

    # ── Show results ───────────────────────────────────────────────────────
    if "results_df" in st.session_state:
        results_df = st.session_state["results_df"]

        st.markdown("---")
        st.markdown("### 📊 Ringkasan Hasil")

        counts = results_df["Status Kesiapan"].value_counts()
        total = len(results_df)
        c1, c2, c3, c4 = st.columns(4)
        for col, key, emoji, color in [
            (c1, "total", "👥", "#a78bfa"),
            (c2, "Siap Olimpiade", "🏆", "#10b981"),
            (c3, "Potensial", "⭐", "#f59e0b"),
            (c4, "Tidak Siap", "📚", "#ef4444"),
        ]:
            val = total if key == "total" else counts.get(key, 0)
            pct = "" if key == "total" else f"({val/total*100:.1f}%)"
            with col:
                st.markdown(f"""<div class="metric-box">
                    <div style="font-size:1.5rem;">{emoji}</div>
                    <div class="metric-val" style="color:{color};">{val}</div>
                    <div class="metric-lbl">{"Total Siswa" if key=="total" else key} {pct}</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### 📋 Tabel Hasil Lengkap")

        # Filter
        filter_status = st.selectbox(
            "Filter berdasarkan status:",
            ["Semua", "Siap Olimpiade", "Potensial", "Tidak Siap"]
        )
        df_show = results_df if filter_status == "Semua" else results_df[results_df["Status Kesiapan"] == filter_status]

        display_cols = ["Nama Siswa", "Skor Kesiapan", "Status Kesiapan", "Keyakinan Model (%)"]
        st.dataframe(df_show[display_cols].reset_index(drop=True), use_container_width=True, height=400)

        # Download
        st.markdown("### 📥 Unduh Hasil")
        excel_buf = export_hasil_excel(results_df)
        st.download_button(
            "📥 Download Semua Hasil (Excel)",
            data=excel_buf,
            file_name="OLYMATH_Hasil_Prediksi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ══════════════════════════════════════════════════════════════════════════════
# HALAMAN TENTANG
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📖 Tentang Sistem":
    st.markdown("## 📖 Tentang OLYMATH")
    st.markdown(f"""
    <div class="card">
        <div class="card-title">🏆 Apa itu OLYMATH?</div>
        <p>OLYMATH (<b>OLYmpiade MATHematics</b>) adalah sistem berbasis Machine Learning untuk membantu 
        guru dan sekolah dalam melakukan seleksi awal calon peserta olimpiade matematika secara objektif, 
        cepat, dan transparan.</p>
    </div>
    <div class="card">
        <div class="card-title">🤖 Metode Klasifikasi</div>
        <p><b>Random Forest Classifier</b> – algoritma ensemble yang menggabungkan banyak decision tree 
        untuk prediksi yang robust dan akurat. Dilatih dengan data 29.000+ siswa dengan akurasi <b>95.4%</b>.</p>
    </div>
    <div class="card">
        <div class="card-title">📊 Variabel Input</div>
        <ul>
            <li><b>Numerasi Aljabar</b> – kemampuan operasi aljabar dan persamaan</li>
            <li><b>Numerasi Geometri</b> – pemahaman bangun ruang dan transformasi</li>
            <li><b>Numerasi Bilangan</b> – operasi bilangan dan teori bilangan</li>
            <li><b>Data & Ketidakpastian</b> – statistika dan peluang</li>
            <li><b>Kemampuan Menalar</b> – logika dan pembuktian matematis</li>
            <li><b>Literasi Matematika</b> – kemampuan membaca dan menginterpretasi soal</li>
        </ul>
    </div>
    <div class="card">
        <div class="card-title">📈 Skor Kesiapan vs Tingkat Keyakinan</div>
        <p><b>Skor Kesiapan</b> adalah indikator utama (0–100) yang selalu naik seiring meningkatnya nilai siswa. 
        Dihitung dengan formula: <code>0.7 × rata-rata + 0.3 × nilai minimum</code> untuk mempertimbangkan 
        konsistensi di semua aspek.</p>
        <p><b>Tingkat Keyakinan Model</b> mencerminkan kepercayaan AI dalam menentukan <i>kategori</i>. 
        Siswa di batas antar-kategori wajar memiliki keyakinan lebih rendah — ini bukan berarti kemampuannya 
        lebih rendah, melainkan posisinya berada di "zona peralihan".</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="card" style="text-align:center; margin-top:20px;">
        <div style="font-size:2rem;">💜</div>
        <div style="color:#a78bfa; font-weight:700; font-size:1.1rem; margin:8px 0;">OLYMATH</div>
        <div style="color:rgba(255,255,255,0.6); font-size:0.85rem;">
            Dibuat untuk Tugas Data Mining | Program Studi Pendidikan Matematika<br>
            Membantu guru & siswa meraih prestasi olimpiade matematika 🏆
        </div>
    </div>
    """, unsafe_allow_html=True)
